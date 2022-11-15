#!/usr/bin/env python2
# -*- coding: utf-8 -*-
""" 
Find FLOPROS protection standards nearest to lon/lat coordinates

Input:
qlon   vector of longitudes to find FLOPROS for
qlat   vector of latitudes to find FLOPROS for

Output:
flopros_near_points   vector of flood protection standards at query locations

Created on Mon Dec 28 11:56:32 2021
@author: Tim Hermans
tim(dot)hermans@nioz(dot)nl
"""
#import fiona
import geopandas as geopd
import pandas as pd
import numpy as np
import os
from shapely.geometry import *
import shapely
from openpyxl import load_workbook

def get_FLOPROS_nearest_to_points(qlon,qlat):

    #load FLOPROS data
    path = '/Users/thermans/Documents/Data/FLOPROS/Tiggeloven/' #path to FLOPROS data (from Tiggeloven et al., 2020)
    polygon_df = geopd.read_file(os.path.join(path,'Results_adaptation_objectives/Countries_States_simplified.shp')) #shape file with region polygons
    wb = load_workbook(filename = os.path.join(path,'FLOPROS_geogunit_107.xlsx')) #protection standards for each region
    ws = wb.active
    flopros = np.array([cell.value for cell in ws['D'][1::]],dtype=np.float)
    
    #put Lon & Lat into pandas df
    locs_df = pd.DataFrame(
        {
         'Latitude': qlat,
         'Longitude': qlon})
    
    locs_df = geopd.GeoDataFrame(
        locs_df, geometry=geopd.points_from_xy(locs_df.Longitude, locs_df.Latitude)) #put input locations into geopd df
        
    dists = polygon_df.geometry.apply(lambda g: locs_df.distance(g)) #calculate distances from all input locations to all (multi)polygons from FLOPROS
    minidx = dists.idxmin() #find indices of polygons with minimum distance to each location
    
    flopros_near_points = flopros[polygon_df.FID_Aque.iloc[minidx]] #output the flopros of these polygons
    
    return flopros_near_points